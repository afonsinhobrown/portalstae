"""
Script para gerar o relatório de análise do projeto em Excel.
Executa: python gerar_relatorio_excel.py
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_FILE = "RELATORIO_ANALISE_PORTALSTAE.xlsx"

wb = Workbook()

# ─── Paleta de Cores ───────────────────────────────────────────────────────────
COR_HEADER     = "1F4E79"   # azul escuro
COR_CRITICO    = "C00000"   # vermelho
COR_MEDIO      = "E36C09"   # laranja
COR_baixo      = "375623"   # verde escuro
COR_OK         = "375623"   # verde
COR_AVISO      = "7F6000"   # amarelo-escuro
COR_ALT1       = "DCE6F1"   # azul claro (linha alternada)
COR_ALT2       = "FFFFFF"   # branco

def header_style(ws, row, cols, title, color=COR_HEADER):
    """Mescla células e aplica estilo de cabeçalho."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color="FFFFFF", size=13)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def col_headers(ws, row, headers, bg=COR_HEADER):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 18

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def data_row(ws, row, values, alt=False):
    bg = COR_ALT1 if alt else COR_ALT2
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()
        c.font = Font(size=10)
    ws.row_dimensions[row].height = 16

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def priority_cell(ws, row, col, priority):
    colors = {"🔴 CRÍTICO": "C00000", "🟡 MÉDIO": "E36C09", "🟢 BAIXO": "375623"}
    fg_map = {"🔴 CRÍTICO": "FFFFFF", "🟡 MÉDIO": "FFFFFF", "🟢 BAIXO": "FFFFFF"}
    c = ws.cell(row=row, column=col, value=priority)
    c.fill = PatternFill("solid", fgColor=colors.get(priority, "999999"))
    c.font = Font(bold=True, color=fg_map.get(priority, "FFFFFF"), size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# Aba 1 — Resumo do Projeto
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Resumo do Projeto"

header_style(ws1, 1, 4, "🏛️  PORTAL STAE — Relatório de Análise de Módulos")
ws1.merge_cells("A2:D2")
ws1.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}").font = Font(italic=True, color="595959", size=9)
ws1.cell(2, 1).alignment = Alignment(horizontal="right")

col_headers(ws1, 4, ["Módulo", "Apps Django", "Nº Modelos", "Estado Geral"])
modulos = [
    ("Recursos Humanos",    "recursoshumanos",        "12+",  "✅ Funcional"),
    ("Gestão Combustível",  "gestaocombustivel",      "15+",  "⚠️ models.py corrompido"),
    ("Gestão Equipamentos", "gestaoequipamentos",     "6",    "✅ Funcional (melhorias)"),
    ("Credenciais",         "credenciais",            "11",   "✅ Funcional"),
    ("Logística Eleitoral", "rs",                     "7",    "✅ Funcional"),
    ("Resultados Eleitorais","dfec",                  "8+",   "✅ Funcional"),
    ("UGEA (Contratos)",    "ugea",                   "5+",   "✅ Funcional"),
    ("Candidaturas",        "candidaturas",           "3+",   "✅ Funcional"),
    ("Partidos",            "partidos",               "2+",   "✅ Funcional"),
    ("Círculos Eleitorais", "circuloseleitorais",     "2+",   "✅ Funcional"),
    ("Eleição",             "eleicao",                "3+",   "✅ Funcional"),
    ("Apuramento",          "apuramento",             "2+",   "✅ Funcional"),
]
for i, row in enumerate(modulos):
    data_row(ws1, i+5, row, alt=i%2==0)
set_col_widths(ws1, [28, 24, 14, 28])


# ══════════════════════════════════════════════════════════════════════════════
# Aba 2 — Gestão de Equipamentos
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📦 Equipamentos")
header_style(ws2, 1, 5, "📦 Módulo: gestaoequipamentos — Análise Detalhada", "1F4E79")
col_headers(ws2, 3, ["Componente", "Tipo", "Detalhe", "Estado", "Observação"])

eq_data = [
    ("CategoriaEquipamento",        "Modelo",   "Agrupa tipos por categoria",                   "✅ OK",        "—"),
    ("TipoEquipamento",             "Modelo",   "Subtipo com atributos JSON",                   "⚠️ Sem schema", "JSONField sem validação"),
    ("Equipamento",                 "Modelo",   "Ativo físico principal",                       "✅ OK",        "—"),
    ("MovimentacaoEquipamento",     "Modelo",   "Tranferência entre setores com aprovação",     "✅ OK",        "—"),
    ("Armazem",                     "Modelo",   "Local de armazenamento",                       "✅ OK",        "—"),
    ("Inventario",                  "Modelo",   "Qtd. de equipamento em armazém",               "❌ Parcial",   "Sem views nem URLs"),
    ("dashboard_equipamentos",      "View",     "KPIs gerais",                                  "✅ OK",        "—"),
    ("lista_equipamentos",          "View",     "Lista com filtros",                            "⚠️ Sem paginação","Lento com muitos registos"),
    ("lista_movimentacoes",         "View",     "Lista de pedidos de movimentação",             "⚠️ Sem paginação","Lento com muitos registos"),
    ("aprovar_movimentacao",        "View",     "Aprovação restrita a is_staff",                "✅ OK",        "—"),
    ("concluir_movimentacao",       "View",     "Atualiza sector_atual do equipamento",         "✅ OK",        "—"),
    ("excluir_equipamento",         "View",     "Soft-delete (em_uso=False)",                   "⚠️ Nome errado","Devia chamar-se desativar_"),
    ("patrimonio_global",           "View",     "Agrega equip. + rs + dfec",                   "⚠️ Import frágil","from dfec.models.completo import..."),
]
for i, row in enumerate(eq_data):
    data_row(ws2, i+4, row, alt=i%2==0)
set_col_widths(ws2, [28, 12, 40, 18, 38])


# ══════════════════════════════════════════════════════════════════════════════
# Aba 3 — Transportes
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("🚗 Transportes")
header_style(ws3, 1, 5, "🚗 Módulo: gestaocombustivel + rs — Análise de Transportes", "1F3864")
col_headers(ws3, 3, ["Modelo/Componente", "App", "Função", "Estado", "Observação"])

transp_data = [
    ("Viatura",                 "gestaocombustivel", "Veículo da frota",                      "✅ OK",         "—"),
    ("PedidoCombustivel",       "gestaocombustivel", "Abastecimento c/ senha de confirmação",  "✅ OK",         "—"),
    ("ManutencaoViatura",       "gestaocombustivel", "Registos de manutenção",                "✅ OK",         "—"),
    ("RotaTransporte",          "gestaocombustivel", "Rotas fixas entre destinos",             "✅ OK",         "—"),
    ("RegistroDiarioRota",      "gestaocombustivel", "Log diário de execução de rota",         "✅ OK",         "—"),
    ("SeguroViatura",           "gestaocombustivel", "Apólices com alerta de vencimento",      "✅ OK",         "—"),
    ("ReservaViatura",          "gestaocombustivel", "Reserva de viatura",                    "❌ DUPLICADA",  "Classe definida 2x no models.py"),
    ("models.py",               "gestaocombustivel", "Ficheiro de modelos",                   "❌ CORROMPIDO", "16.214 linhas, maioria em branco"),
    ("PlanoLogistico",          "rs",                "Plano de logística eleitoral",           "✅ OK",         "—"),
    ("MaterialEleitoral",       "rs",                "Necessidades de materiais p/ eleição",   "✅ OK",         "Ligado a gestaoequipamentos"),
    ("dashboard_combustivel",   "gestaocombustivel", "KPIs e contratos UGEA integrados",       "✅ OK",         "—"),
    ("pedir_combustivel",       "gestaocombustivel", "Pedido com integração UGEA",             "✅ OK",         "—"),
    ("aprovar_pedido",          "gestaocombustivel", "Aprovação e confirmação com senha",      "✅ OK",         "—"),
]
for i, row in enumerate(transp_data):
    data_row(ws3, i+4, row, alt=i%2==0)
set_col_widths(ws3, [26, 22, 38, 16, 36])


# ══════════════════════════════════════════════════════════════════════════════
# Aba 4 — Plano de Ação
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("🔧 Plano de Ação")
header_style(ws4, 1, 5, "🔧 Plano de Ação — Problemas & Correções Recomendadas", "833C00")
col_headers(ws4, 3, ["Prioridade", "Módulo", "Problema", "Ação Recomendada", "Esforço"])

acoes = [
    ("🔴 CRÍTICO", "gestaocombustivel", "ReservaViatura definida 2x",                     "Limpar models.py e remover duplicação",               "Alto"),
    ("🔴 CRÍTICO", "gestaocombustivel", "models.py com 16k linhas em branco",             "Reformatar ficheiro e remover linhas vazias",          "Alto"),
    ("🟡 MÉDIO",   "gestaoequipamentos","Inventario sem views nem URLs",                  "Criar CRUD de inventário com paginação",               "Médio"),
    ("🟡 MÉDIO",   "gestaoequipamentos","Listas sem paginação",                           "Adicionar Paginator em lista_equipamentos e lista_mov","Baixo"),
    ("🟡 MÉDIO",   "IDE (Pyre2)",       "search roots vazios (erros de import)",          ".pyre_configuration criado — recarregar IDE",          "Baixo"),
    ("🟢 BAIXO",   "gestaoequipamentos","import frágil em patrimonio_global",             "Usar from dfec.models import ...",                    "Muito Baixo"),
    ("🟢 BAIXO",   "gestaoequipamentos","excluir_equipamento com nome errado",            "Renomear para desativar_equipamento",                  "Muito Baixo"),
    ("🟢 BAIXO",   "gestaoequipamentos","JSONField sem schema em atributos_especificos",  "Documentar ou validar estrutura do JSON",              "Baixo"),
]

for i, row in enumerate(acoes):
    r = i + 4
    priority_cell(ws4, r, 1, row[0])
    for col, val in enumerate(row[1:], 2):
        c = ws4.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=COR_ALT1 if i%2==0 else COR_ALT2)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = thin_border()
        c.font = Font(size=10)
    ws4.row_dimensions[r].height = 18

set_col_widths(ws4, [16, 22, 38, 44, 12])


# ─── Guardar ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"[OK] Ficheiro Excel criado com sucesso: {OUTPUT_FILE}")
print(f"     Localizacao: {__import__('os').path.abspath(OUTPUT_FILE)}")
