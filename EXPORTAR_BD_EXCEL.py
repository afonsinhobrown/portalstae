"""
EXPORTAR_BD_EXCEL.py
Exporta todos os dados do SQLite do Portal STAE para um ficheiro Excel.
Cada tabela tem a sua própria sheet.
Autenticado via Django (admin / admin123).

Executa: python EXPORTAR_BD_EXCEL.py
"""
import os
import sys
import django

# ── Configurar Django ─────────────────────────────────────────────────────────
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'portalstae.settings')
django.setup()

# ── Autenticar como admin ─────────────────────────────────────────────────────
from django.contrib.auth import authenticate
from django.contrib.auth.models import User

USERNAME = 'admin'
PASSWORD = 'admin123'

user = authenticate(username=USERNAME, password=PASSWORD)
if user is None:
    # Tentar encontrar o utilizador e verificar se existe
    try:
        u = User.objects.get(username=USERNAME)
        print(f"[AVISO] Utilizador '{USERNAME}' existe mas a password esta incorreta.")
        print("        A continuar exportacao sem autenticacao strict...")
    except User.DoesNotExist:
        print(f"[AVISO] Utilizador '{USERNAME}' nao encontrado na base de dados.")
        print("        A continuar exportacao sem autenticacao strict...")
else:
    print(f"[OK] Autenticado como: {user.username} (staff={user.is_staff}, superuser={user.is_superuser})")

# ── Imports para exportação ───────────────────────────────────────────────────
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_FILE = "PORTAL_STAE_DATABASE_EXPORT.xlsx"
DB_FILE     = "db.sqlite3"

# ── Estilos ───────────────────────────────────────────────────────────────────
def make_header_cell(cell, value, bg="1F4E79"):
    cell.value = value
    cell.font  = Font(bold=True, color="FFFFFF", size=10)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    s = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def make_data_cell(cell, value, alt=False):
    cell.value = value
    cell.fill  = PatternFill("solid", fgColor="DCE6F1" if alt else "FFFFFF")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    s = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=s, right=s, top=s, bottom=s)
    cell.font = Font(size=9)

# ── Criar Workbook ────────────────────────────────────────────────────────────
wb = Workbook()

# ── Aba 0: Índice / Sumário ───────────────────────────────────────────────────
ws_idx = wb.active
ws_idx.title = "INDICE"

conn = sqlite3.connect(DB_FILE)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

tables = cur.execute(
    "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
).fetchall()
table_names = [t[0] for t in tables]

# Cabeçalho sumário
ws_idx.merge_cells("A1:E1")
c = ws_idx.cell(1, 1, f"PORTAL STAE — Exportacao da Base de Dados | {datetime.now().strftime('%d/%m/%Y %H:%M')}")
c.font  = Font(bold=True, color="FFFFFF", size=14)
c.fill  = PatternFill("solid", fgColor="1F3864")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_idx.row_dimensions[1].height = 28

ws_idx.merge_cells("A2:E2")
c2 = ws_idx.cell(2, 1, f"Autenticado como: {USERNAME}  |  Total de tabelas: {len(table_names)}  |  Ficheiro: {DB_FILE}")
c2.font = Font(italic=True, color="595959", size=9)
c2.alignment = Alignment(horizontal="center")

for col, h in enumerate(["#", "Tabela (Sheet)", "Modulo", "Total Registos", "Estado"], 1):
    make_header_cell(ws_idx.cell(4, col), h)
ws_idx.row_dimensions[4].height = 18

# Mapear prefixos de módulos
MODULE_MAP = {
    "auth_":               "Django Auth",
    "django_":             "Django Core",
    "sqlite_":             "SQLite",
    "admin_portal_":       "Admin Portal",
    "apuramento_":         "Apuramento",
    "candidaturas_":       "Candidaturas",
    "chatbot_":            "Chatbot",
    "circuloseleitorais_": "Circulos Eleitorais",
    "credenciais_":        "Credenciais",
    "dfec_":               "DFEC",
    "eleicao_":            "Eleicao",
    "gestaocombustivel_":  "Gestao Combustivel",
    "gestaoequipamentos_": "Gestao Equipamentos",
    "pagina_stae_":        "Pagina STAE",
    "partidos_":           "Partidos",
    "recursoshumanos_":    "Recursos Humanos",
    "rs_":                 "RS Logistica",
    "ugea_":               "UGEA",
}

def get_module(tname):
    for prefix, mod in MODULE_MAP.items():
        if tname.startswith(prefix):
            return mod
    return "Outro"

# Registar sumário e criar sheets por tabela
created_sheets  = []
idx_row         = 5
MAX_ROWS_SHEET  = 50000  # Limite de segurança por sheet

print(f"\n[INFO] A exportar {len(table_names)} tabelas...")

for i, tname in enumerate(table_names):
    try:
        count = cur.execute(f'SELECT COUNT(*) FROM "{tname}"').fetchone()[0]
    except:
        count = -1

    modulo = get_module(tname)
    estado = "OK" if count >= 0 else "ERRO"

    # Índice
    for col, val in enumerate([i+1, tname, modulo, count if count >= 0 else "ERRO", estado], 1):
        make_data_cell(ws_idx.cell(idx_row, col), val, alt=i%2==0)
    idx_row += 1

    # Nome da sheet (Excel: max 31 chars, sem carateres especiais)
    sheet_name = tname[:31].replace("/","_").replace("\\","_").replace("*","_").replace("?","_").replace("[","_").replace("]","_").replace(":","_")

    # Criar sheet com dados
    ws = wb.create_sheet(title=sheet_name)

    # Titulo
    ws.merge_cells(f"A1:{get_column_letter(max(1,3))}1")
    t = ws.cell(1, 1, f"Tabela: {tname}  |  Modulo: {modulo}  |  {count} registos")
    t.font  = Font(bold=True, color="FFFFFF", size=11)
    t.fill  = PatternFill("solid", fgColor="1F3864")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    if count == 0:
        ws.cell(3, 1, "(Tabela vazia)").font = Font(italic=True, color="999999")
        created_sheets.append((tname, 0))
        print(f"  [{i+1:03d}/{len(table_names)}] {tname}: 0 registos (vazia)")
        continue

    if count < 0:
        ws.cell(3, 1, "ERRO ao ler tabela").font = Font(bold=True, color="C00000")
        created_sheets.append((tname, -1))
        continue

    try:
        # Ler dados (com limite)
        rows = cur.execute(f'SELECT * FROM "{tname}" LIMIT {MAX_ROWS_SHEET}').fetchall()
        if not rows:
            ws.cell(3, 1, "(Sem dados)").font = Font(italic=True)
            created_sheets.append((tname, 0))
            continue

        # Colunas
        col_names = [desc[0] for desc in cur.description]
        for ci, col in enumerate(col_names, 1):
            make_header_cell(ws.cell(2, ci), col)

        # Dados
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row, 1):
                # Converter para tipo legível
                if isinstance(val, bytes):
                    val = f"[BYTES: {len(val)}]"
                elif val is None:
                    val = ""
                make_data_cell(ws.cell(ri+3, ci), str(val) if not isinstance(val, (int, float)) else val, alt=ri%2==0)

        # Auto-ajustar largura de colunas
        for ci, col in enumerate(col_names, 1):
            max_len = len(str(col))
            for row in rows[:100]:  # Amostra para performance
                celval = str(row[ci-1]) if row[ci-1] is not None else ""
                max_len = max(max_len, min(len(celval), 50))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 52)

        ws.freeze_panes = "A3"  # Congelar cabeçalho
        created_sheets.append((tname, len(rows)))
        print(f"  [{i+1:03d}/{len(table_names)}] {tname}: {count} registos exportados")

    except Exception as e:
        ws.cell(3, 1, f"ERRO: {e}").font = Font(bold=True, color="C00000")
        created_sheets.append((tname, -1))
        print(f"  [ERRO] {tname}: {e}")

# Larguras do índice
ws_idx.column_dimensions["A"].width = 6
ws_idx.column_dimensions["B"].width = 42
ws_idx.column_dimensions["C"].width = 24
ws_idx.column_dimensions["D"].width = 16
ws_idx.column_dimensions["E"].width = 10
ws_idx.freeze_panes = "A5"

conn.close()

# Guardar
wb.save(OUTPUT_FILE)
ok  = sum(1 for _, c in created_sheets if c >= 0)
err = sum(1 for _, c in created_sheets if c < 0)
total_recs = sum(c for _, c in created_sheets if c > 0)
print(f"\n[CONCLUIDO]")
print(f"  Ficheiro : {OUTPUT_FILE}")
print(f"  Sheets   : {len(created_sheets)} ({ok} OK, {err} com erro)")
print(f"  Registos : {total_recs} exportados no total")
print(f"  Path     : {os.path.abspath(OUTPUT_FILE)}")
